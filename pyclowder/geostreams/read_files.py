import datetime
import logging

from openpyxl import load_workbook


def read_xlsx(file_path):
    logging.info("[pyG.read_files] Loading " + str(file_path))
    wb = load_workbook(file_path, data_only=True, read_only=True)
    sheet_names = wb.get_sheet_names()
    sheet = wb.get_sheet_by_name(sheet_names[0])
    logging.info("[pyG.read_files] Done loading " + str(file_path))

    ### GET THE COLUMN NAMES WHICH ARE LOCATED IN THE FIRST ROW ########################################
    logging.info("[pyG.read_files] getting column headers")
    column_names = []
    i = 1
    j = 1
    while (1):
        value = sheet.cell(row=i, column=j).value
        if value == None:  ### READ ROW UNTIL CELL IS EMPTY -> RETURNS None
            break
        column_names.append(value)
        j += 1

    num_cols = len(column_names)

    ### GET THE FIELD VALUES WHICH START IN THE SECOND ROW ########################################
    logging.info("[pyG.read_files] loading data into list")
    field_values = []
    i = 2
    count = 0
    while (1):
        count += 1
        # if 500%count==0:
        logging.info("[pyG.read_files] Loaded data row " + str(count))
        j = 1
        try:
            value = sheet.cell(row=i, column=j).value
        except:
            break
        field_row = []
        while (1):
            if j == num_cols + 1:
                break
            value = sheet.cell(row=i, column=j).value
            if isinstance(value, datetime.date) or isinstance(value, datetime.time):
                value = str(value)
            field_row.append(value)
            j += 1
        if all(v is None for v in field_row):
            break
        field_values.append(field_row)
        i += 1
    logging.info("[pyG.read_files] Done loading data into list")
    return column_names, field_values


def read_tsv(file_path):
    input_file = open(file_path, 'r')
    raw_text = input_file.read()

    ### READ DATA INTO ROWS EXCLUDING LINES STARTING WITH '#'
    rows = []
    line_text = ""

    for i in range(len(raw_text)):
        if raw_text[i] != "\n":

            line_text += raw_text[i]
        else:
            if line_text[0] != "#":
                rows.append(line_text)
            line_text = ""

    headers = rows.pop(0).split("\t")

    logging.info("[pyG.read_files] Creating 2D list of data")
    data = []
    while len(rows) > 0:
        data.append(rows.pop(0).split("\t"))
    del raw_text
    del rows
    logging.info("[pyG.read_files] Done Creating 2D list of data")

    return headers, data


def text2list_noaa(response_in):
    data_out = []
    column_names = []
    header_names = []
    for key in response_in:
        rows = []
        data_site = []
        lineOfText = ""
        for i in range(len(response_in[key])):
            if response_in[key][i] != "\n":
                lineOfText += response_in[key][i]
            else:
                rows.append(lineOfText)
                lineOfText = ""
        header_text = rows.pop(0)
        header_names = header_text.split("\t")
        header_names.pop(1)
        for i in range(len(rows)):
            data_site.append(rows[i].split("\t"))
            station_text = data_site[i][0].split(":")
            data_site[i][0] = station_text[len(station_text) - 1]
            sensor_text = data_site[i][1].split(":")
            data_site[i][1] = sensor_text[len(sensor_text) - 2]

            if data_site[i][0] != data_site[i][1]:
                print "Station ID and Sensor ID do not match"
                quit()
            else:
                data_site[i].pop(0)

            # added for water level which has a datum_id for the reference water level
            if len(header_names) > 5:
                if header_names[5] == "datum_id":
                    datum_id_text = data_site[i][5].split(":")
                    data_site[i][5] = datum_id_text[len(datum_id_text) - 1]

        header_names[0] = u"station_id"

    # TODO: as is, this only processes 1 parameter, needs to be changed to put multiple paratmeters into data_out

    return data_site, header_names